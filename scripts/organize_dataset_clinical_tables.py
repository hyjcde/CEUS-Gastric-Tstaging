#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATASET_ROOT = PROJECT_ROOT / "dataset"
OUTPUT_ROOT = DATASET_ROOT / "tables"
RAW_COPY_ROOT = OUTPUT_ROOT / "raw"
BY_SOURCE_ROOT = OUTPUT_ROOT / "by_source"


@dataclass(frozen=True)
class TableSource:
    table_id: str
    cohort: str
    center: str
    year_group: str
    workbook_path: Path


EXTERNAL_SOURCE_ROOT = PROJECT_ROOT / "胃癌直接手术外部测试集" / "直接手术图片"

SOURCES: list[TableSource] = [
    TableSource(
        table_id="internal_2018_direct_surgery",
        cohort="internal_training",
        center="协和内部",
        year_group="2018",
        workbook_path=PROJECT_ROOT / "胃癌分期/协和内部数据集/直接手术/表格整理/2018直接手术.xlsx",
    ),
    TableSource(
        table_id="internal_2019_direct_surgery",
        cohort="internal_training",
        center="协和内部",
        year_group="2019",
        workbook_path=PROJECT_ROOT / "胃癌分期/协和内部数据集/直接手术/表格整理/2019.xlsx",
    ),
    TableSource(
        table_id="internal_2020_2023_direct_surgery",
        cohort="internal_training",
        center="协和内部",
        year_group="2020_2023",
        workbook_path=PROJECT_ROOT / "胃癌分期/协和内部数据集/直接手术/表格整理/20-23年直接手术胃癌.xlsx",
    ),
    TableSource(
        table_id="internal_2024_direct_surgery",
        cohort="internal_training",
        center="协和内部",
        year_group="2024",
        workbook_path=PROJECT_ROOT / "胃癌分期/协和内部数据集/直接手术/表格整理/2024年胃癌手术.xlsx",
    ),
    TableSource(
        table_id="internal_2025_direct_surgery",
        cohort="internal_prospective",
        center="协和内部",
        year_group="2025",
        workbook_path=PROJECT_ROOT / "胃癌分期/协和内部数据集/直接手术/表格整理/2025胃癌直接手术.xlsx",
    ),
    TableSource(
        table_id="external_sanming_direct_surgery",
        cohort="external_test",
        center="三明市第二医院",
        year_group="external",
        workbook_path=EXTERNAL_SOURCE_ROOT / "三明市第二医院" / "三明胃癌临床资料.xlsx",
    ),
    TableSource(
        table_id="external_tumor_hospital_direct_surgery",
        cohort="external_test",
        center="福建省肿瘤医院",
        year_group="external",
        workbook_path=EXTERNAL_SOURCE_ROOT / "福建省肿瘤医院" / "肿瘤医院临床资料.xlsx",
    ),
    TableSource(
        table_id="external_putian1_direct_surgery",
        cohort="external_test",
        center="莆田学院附属医院",
        year_group="external",
        workbook_path=EXTERNAL_SOURCE_ROOT / "莆田学院附属医院" / "莆田1胃癌临床数据.xlsx",
    ),
    TableSource(
        table_id="external_putian2_direct_surgery",
        cohort="external_test",
        center="莆田市第一医院",
        year_group="external",
        workbook_path=EXTERNAL_SOURCE_ROOT / "莆田市第一医院" / "莆田2胃癌临床资料.xlsx",
    ),
    TableSource(
        table_id="external_beijing_friendship_direct_surgery",
        cohort="external_test",
        center="北京友谊医院",
        year_group="external",
        workbook_path=EXTERNAL_SOURCE_ROOT / "北京友谊医院" / "胃癌临床资料.xlsx",
    ),
    TableSource(
        table_id="external_foshan_first_direct_surgery",
        cohort="external_test",
        center="佛山市第一人民医院",
        year_group="external",
        workbook_path=EXTERNAL_SOURCE_ROOT / "佛山市第一人民医院" / "广东胃癌临床资料模板.xlsx",
    ),
    TableSource(
        table_id="external_cnnc_504_direct_surgery",
        cohort="external_test",
        center="中核五〇四医院",
        year_group="external",
        workbook_path=EXTERNAL_SOURCE_ROOT / "中核五O四医院" / "湖北胃癌临床资料模板.xlsx",
    ),
    TableSource(
        table_id="external_dehua_direct_surgery",
        cohort="external_test",
        center="福建省德化县医院",
        year_group="external",
        workbook_path=EXTERNAL_SOURCE_ROOT / "福建省德化县医院" / "德化胃癌临床资料.xlsx",
    ),
    TableSource(
        table_id="external_fujian_provincial_direct_surgery",
        cohort="external_test",
        center="福建省立医院",
        year_group="external",
        workbook_path=EXTERNAL_SOURCE_ROOT / "福建省立医院" / "胃癌临床资料模板.xlsx",
    ),
]


CANONICAL_FIELDS = [
    "record_key_raw",
    "record_key_type",
    "patient_name_raw",
    "sex_code_raw",
    "age_raw",
    "long_diameter_cm_raw",
    "thickness_cm_raw",
    "tumor_location_raw",
    "cea_raw",
    "cea_positive_raw",
    "ca199_raw",
    "ca199_positive_raw",
    "pathology_raw",
    "differentiation_raw",
    "lauren_raw",
    "t_stage_raw",
    "ultrasound_exam_date_raw",
]


def ensure_clean_dir(path: Path) -> Path:
    if path.exists():
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)
    return path


def stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def normalize_header(text: str) -> str:
    return text.replace("\n", "").replace(" ", "").replace("\u3000", "")


def unique_headers(values: Iterable[object]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for idx, value in enumerate(values, start=1):
        base = stringify(value) or f"column_{idx}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def extract_canonical_fields(row_map: dict[str, str]) -> dict[str, str]:
    canonical = {field: "" for field in CANONICAL_FIELDS}

    for header, value in row_map.items():
        if not value:
            continue

        key = normalize_header(header)
        if key == "住院号":
            canonical["record_key_raw"] = value
            canonical["record_key_type"] = "hospitalization_no"
        elif key == "ID":
            canonical["record_key_raw"] = value
            canonical["record_key_type"] = "table_id"
        elif key == "序号":
            canonical["record_key_raw"] = value
            canonical["record_key_type"] = "table_sequence"
        elif key == "姓名":
            canonical["patient_name_raw"] = value
        elif key.startswith("性别"):
            canonical["sex_code_raw"] = value
        elif key.startswith("年龄"):
            canonical["age_raw"] = value
        elif key.startswith("长径"):
            canonical["long_diameter_cm_raw"] = value
        elif key.startswith("厚径"):
            canonical["thickness_cm_raw"] = value
        elif "肿瘤位置" in key or "超声位置" in key:
            canonical["tumor_location_raw"] = value
        elif key == "CEA":
            canonical["cea_raw"] = value
        elif key.startswith("CEA：") and "阴性" in key:
            canonical["cea_positive_raw"] = value
        elif (key == "CA199" or key == "CA19-9") and "阴性" not in key:
            canonical["ca199_raw"] = value
        elif (key.startswith("CA199：") or key.startswith("CA199:") or key.startswith("CA19-9")) and "阴性" in key:
            canonical["ca199_positive_raw"] = value
        elif "病理" in key:
            canonical["pathology_raw"] = value
        elif "分化程度" in key:
            canonical["differentiation_raw"] = value
        elif "Lauren" in key:
            canonical["lauren_raw"] = value
        elif key.startswith("T:"):
            canonical["t_stage_raw"] = value
        elif "超声检查时间" in key:
            canonical["ultrasound_exam_date_raw"] = value

    return canonical


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    RAW_COPY_ROOT.mkdir(parents=True, exist_ok=True)
    BY_SOURCE_ROOT.mkdir(parents=True, exist_ok=True)
    ensure_clean_dir(RAW_COPY_ROOT)
    ensure_clean_dir(BY_SOURCE_ROOT)

    index_rows: list[dict[str, str]] = []
    registry_rows: list[dict[str, str]] = []

    for source in SOURCES:
        if not source.workbook_path.exists():
            raise FileNotFoundError(f"Missing workbook: {source.workbook_path}")

        raw_copy_path = RAW_COPY_ROOT / f"{source.table_id}.xlsx"
        shutil.copy2(source.workbook_path, raw_copy_path)

        workbook = load_workbook(source.workbook_path, read_only=True, data_only=True)
        for sheet_idx, sheet_name in enumerate(workbook.sheetnames, start=1):
            worksheet = workbook[sheet_name]
            nonempty_rows = [
                [stringify(cell) for cell in row]
                for row in worksheet.iter_rows(values_only=True)
                if any(stringify(cell) for cell in row)
            ]
            if len(nonempty_rows) < 2:
                continue

            headers = unique_headers(nonempty_rows[0])
            source_rows: list[dict[str, str]] = []
            for row_offset, values in enumerate(nonempty_rows[1:], start=2):
                padded_values = list(values) + [""] * max(0, len(headers) - len(values))
                row_map = {headers[idx]: padded_values[idx] for idx in range(len(headers))}

                source_record = {
                    "table_id": source.table_id,
                    "cohort": source.cohort,
                    "center": source.center,
                    "year_group": source.year_group,
                    "sheet_name": sheet_name,
                    "source_row_number": str(row_offset),
                    **row_map,
                }
                source_rows.append(source_record)

                registry_rows.append(
                    {
                        "table_id": source.table_id,
                        "cohort": source.cohort,
                        "center": source.center,
                        "year_group": source.year_group,
                        "sheet_name": sheet_name,
                        "source_row_number": str(row_offset),
                        "source_workbook_path": str(source.workbook_path.relative_to(PROJECT_ROOT)),
                        **extract_canonical_fields(row_map),
                        "source_row_json": json.dumps(row_map, ensure_ascii=False),
                    }
                )

            export_name = f"{source.table_id}__sheet{sheet_idx}.csv"
            export_path = BY_SOURCE_ROOT / export_name
            write_csv(
                export_path,
                source_rows,
                [
                    "table_id",
                    "cohort",
                    "center",
                    "year_group",
                    "sheet_name",
                    "source_row_number",
                    *headers,
                ],
            )

            index_rows.append(
                {
                    "table_id": source.table_id,
                    "cohort": source.cohort,
                    "center": source.center,
                    "year_group": source.year_group,
                    "sheet_name": sheet_name,
                    "row_count": str(len(source_rows)),
                    "source_workbook_path": str(source.workbook_path.relative_to(PROJECT_ROOT)),
                    "raw_copy_path": str(raw_copy_path.relative_to(PROJECT_ROOT)),
                    "export_csv_path": str(export_path.relative_to(PROJECT_ROOT)),
                    "source_headers_json": json.dumps(headers, ensure_ascii=False),
                }
            )

    write_csv(
        OUTPUT_ROOT / "clinical_table_index.csv",
        index_rows,
        [
            "table_id",
            "cohort",
            "center",
            "year_group",
            "sheet_name",
            "row_count",
            "source_workbook_path",
            "raw_copy_path",
            "export_csv_path",
            "source_headers_json",
        ],
    )
    write_csv(
        OUTPUT_ROOT / "clinical_table_registry.csv",
        registry_rows,
        [
            "table_id",
            "cohort",
            "center",
            "year_group",
            "sheet_name",
            "source_row_number",
            "source_workbook_path",
            *CANONICAL_FIELDS,
            "source_row_json",
        ],
    )

    summary = {
        "output_root": str(OUTPUT_ROOT),
        "table_file_count": len(index_rows),
        "registry_row_count": len(registry_rows),
    }
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
